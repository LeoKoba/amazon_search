from jmespath import search
from selenium import webdriver
import time
from openpyxl import Workbook, load_workbook
from datetime import date
from zipfile import ZipFile


def search_amazon(driver, product):
    ''' Receives the content to be searched and returns two lists with the name and price of the products '''

    name_list = []
    price_list = []

    search_bar = driver.find_element("id", "twotabsearchtextbox")
    search_bar.send_keys(product)

    send_button = driver.find_element("id","nav-search-submit-button")
    send_button.click()

    time.sleep(2)

    name_results = driver.find_elements("xpath", '//span[@class = "a-size-base a-color-base a-text-normal"]')
    price_results = driver.find_elements("xpath", '//span[@class = "a-price-whole"]')

    for i in range(len(name_results)):
        name_list.append(name_results[i].text)
        try:
            price_list.append(price_results[i].text)
        except:
            price_list.append("NO VALUE")

    return name_list, price_list

def list_excel(driver, name_list, price_list):
    ''' Read the lists, fill the file and save with the current date '''
    wb = Workbook()
    ws = wb.active

    ws.title = 'Iphone List'
    ws['A1']= 'Product name'
    ws['B1']= 'Price name'
    
    for i in range(len(price_list)):
        ws.append([name_list[i], price_list[i]])
        wb.save("IPhone_Amazon_{actual_date}.xlsx".format(actual_date = date.today()))

def main():
    ''' main function, where parameters are passed and functions are called '''
    driver = webdriver.Firefox()
    driver.get("https://www.amazon.com.br/")

    name_list, price_list = search_amazon(driver, "IPhone")
    list_excel(driver, name_list, price_list)

if __name__ == '__main__':
   main()


